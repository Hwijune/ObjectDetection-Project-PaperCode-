
# coding: utf-8

# # Object Detection Demo
# Welcome to the object detection inference walkthrough!  This notebook will walk you step by step through the process of using a pre-trained model to detect objects in an image. Make sure to follow the [installation instructions](https://github.com/tensorflow/models/blob/master/object_detection/g3doc/installation.md) before you start.

# # Imports

# In[ ]:


import numpy as np
import os
import six.moves.urllib as urllib
import sys
import tarfile
import tensorflow as tf
import zipfile
import time
import copy
import pymysql

from collections import defaultdict
from io import StringIO
from matplotlib import pyplot as plt
from PIL import Image
import math

import cv2
imgfile = 'test_images/img1.jpg'

# ## Env setup

# In[ ]:

# This is needed since the notebook is stored in the object_detection folder.
sys.path.append("..")


# ## Object detection imports
# Here are the imports from the object detection module.

# In[ ]:


from utils import label_map_util

from utils import visualization_utils as vis_util


# # Model preparation

# ## Variables
#
# Any model exported using the `export_inference_graph.py` tool can be loaded here simply by changing `PATH_TO_CKPT` to point to a new .pb file.
#
# By default we use an "SSD with Mobilenet" model here. See the [detection model zoo](https://github.com/tensorflow/models/blob/master/object_detection/g3doc/detection_model_zoo.md) for a list of other models that can be run out-of-the-box with varying speeds and accuracies.

# In[ ]:
import threading, requests, time


class Thread(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        # What model to download.
        MODEL_NAME = 'ssd_mobilenet_v1_ppn_shared_box_predictor_300x300_coco14_sync_2018_07_03'
        MODEL_FILE = MODEL_NAME + '.tar.gz'
        DOWNLOAD_BASE = 'http://download.tensorflow.org/models/object_detection/'

        # Path to frozen detection graph. This is the actual model that is used for the object detection.
        PATH_TO_CKPT = MODEL_NAME + '/frozen_inference_graph.pb'

        # List of the strings that is used to add correct label for each box.
        PATH_TO_LABELS = os.path.join('data', 'mscoco_label_map.pbtxt')

        NUM_CLASSES = 90


        # ## Download Model

        # In[ ]:


        opener = urllib.request.URLopener()
        opener.retrieve(DOWNLOAD_BASE + MODEL_FILE, MODEL_FILE)
        tar_file = tarfile.open(MODEL_FILE)
        for file in tar_file.getmembers():
          file_name = os.path.basename(file.name)
          if 'frozen_inference_graph.pb' in file_name:
            tar_file.extract(file, os.getcwd())


        # ## Load a (frozen) Tensorflow model into memory.

        # In[ ]:


        detection_graph = tf.Graph()
        with detection_graph.as_default():
          od_graph_def = tf.GraphDef()
          with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
            serialized_graph = fid.read()
            od_graph_def.ParseFromString(serialized_graph)
            tf.import_graph_def(od_graph_def, name='')


        # ## Loading label map
        # Label maps map indices to category names, so that when our convolution network predicts `5`, we know that this corresponds to `airplane`.  Here we use internal utility functions, but anything that returns a dictionary mapping integers to appropriate string labels would be fine

        # In[ ]:


        label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
        categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
        category_index = label_map_util.create_category_index(categories)


        # ## Helper code

        # In[ ]:


        def load_image_into_numpy_array(image):
          (im_width, im_height) = image.size
          return np.array(image.getdata()).reshape(
              (im_height, im_width, 3)).astype(np.uint8)


        def hconcat_resize_min(im_list, interpolation=cv2.INTER_CUBIC):
            h_min = min(im.shape[0] for im in im_list)
            im_list_resize = [
                cv2.resize(im, (int(im.shape[1] * h_min / im.shape[0]), h_min), interpolation=interpolation)
                for im in im_list]
            return cv2.hconcat(im_list_resize)

        # # Detection

        # In[ ]:


        # For the sake of simplicity we will use only 2 images:
        # image1.jpg
        # image2.jpg
        # If you want to test the code with your images, just add path to the images to the TEST_IMAGE_PATHS.
        PATH_TO_TEST_IMAGES_DIR = 'test_images'
        TEST_IMAGE_PATHS = [ os.path.join(PATH_TO_TEST_IMAGES_DIR, 'img{}.jpg'.format(i)) for i in range(1, 3) ]

        # Size, in inches, of the output images.
        IMAGE_SIZE = (12, 8)

        #---------------------------------------------- matplot
        #그래프 그리기
        #그래프 값
        fsize = 400
        framesize = 0
        framesize2 = 0
        frame = [0]
        frame2 = [0]
        x = [0]
        y = [0]
        sumlist = [0] ##합산값 배열
        scorelist = [0] #스코어배열
        ratiolist = [0] #비율배열
        degreelist = [0] #각도 배열
        others = [0] #기타 물체배열

        plt.ion() #그래프창 띄워놓기
        fig = plt.figure(figsize=(4,10))

        sf = fig.add_subplot(511) #3행 1열 1
        sf.set_xlabel("x+y")
        plt.xlim(0,fsize) #그래프 처음끝값
        plt.ylim(0,1500)
        plt.grid(True)
        line1, =  sf.plot(x,y,'r-') #빨강점선 sum값

        sf2 = fig.add_subplot(512) #3행 1열 2
        sf2.set_xlabel("score")
        plt.xlim(0,fsize) #그래프 처음끝값
        plt.ylim(50,105)
        plt.grid(True)
        line2, =  sf2.plot(x,y,'b-') #파랑점선 정확도값

        sf3 = fig.add_subplot(513) #3행 1열 3
        sf3.set_xlabel("(width/height)ratio")
        plt.xlim(0,fsize) #그래프 처음끝값
        plt.ylim(0,5)
        plt.grid(True)
        line3, =  sf3.plot(x,y,'g-') #초록점선 비율값

        sf5 = fig.add_subplot(514) #4행
        sf5.set_xlabel("ellipse degree")
        plt.xlim(0,fsize) #그래프 처음끝값
        plt.ylim(0,90)
        plt.grid(True)
        line4, =  sf5.plot(x,y,'c-') #초록점선 비율값

        sf4 = fig.add_subplot(515)
        sf4.set_xlabel("others")
        plt.xlim(0,fsize*5) #그래프 처음끝값
        plt.ylim(100,1500)
        plt.grid(True)
        dots  = sf4.scatter(x,y) #다른것 점찍기

        #---------------------- excel,mysql

        from openpyxl import Workbook

        wb = Workbook() #워크북 생성
        ws = wb.active #워크 시트 얻기

        # MySQL Connection 연결
        #conn = pymysql.connect(host='localhost', user='root', password='rootpass',
        #                       db='testdb', charset='utf8')

        # Connection 으로부터 Cursor 생성
        #curs = conn.cursor()

        #----------------------- opencv
        filename = 'test1'
        cap = cv2.VideoCapture('101.mp4')#'experiment/'+filename+'.avi'
        fourcc = cv2.VideoWriter_fourcc(*'XVID')
        out = cv2.VideoWriter(filename+'_out_rcnninception.avi', fourcc, 20.0, (960, 720))

        tt, bg = copy.deepcopy(cap.read()) #초기샷
        #bg = cv2.imread('test_images/bg.png', cv2.IMREAD_COLOR)
        bg = cv2.resize(bg, (480, 360))
        bg = cv2.cvtColor(bg, cv2.COLOR_BGR2GRAY)
        bg = copy.deepcopy(cv2.cvtColor(bg, cv2.COLOR_GRAY2RGB)) #흑백으로 바꿨다 컬러로
        backbg = copy.deepcopy(bg)
        countsecond = 1 #초세기

        #detection
        Detectflag = True # 디텍션 이프문 처음에만 들어가기
        # 영상 블러처리 -히트맵
        start = 1
        duration = 10
        fps = '30'

        _, f = cap.read()
        f = cv2.cvtColor(f, cv2.COLOR_BGR2GRAY)
        f = cv2.GaussianBlur(f, (11, 11), 2, 2)
        res = 0.05 * f
        res = res.astype(np.float64)

        fgbg = cv2.createBackgroundSubtractorMOG2(history=1, varThreshold=100,
                                                  detectShadows=True)

        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (13, 13))

        # sql = """insert into testtable(people)
        #          values (%s)"""
        threadtime = 0.5  # 스레드시간
        threadtimesum = 0  # 스레드 시간 합
        with detection_graph.as_default():
          with tf.Session(graph=detection_graph) as sess:
            while True:
                  time.sleep(threadtime)
                  threadtimesum += threadtime

                  ret, image_np = cap.read()

                  image_np3 = copy.deepcopy(image_np)  # 1번화면 히트맵에 복사

                  # 히트맵 그리기
                  fgmask = fgbg.apply(image_np3, None, 0.01)
                  gray = cv2.cvtColor(image_np3, cv2.COLOR_BGR2GRAY)
                  gray = cv2.GaussianBlur(gray, (11, 11), 2, 2)
                  gray = gray.astype(np.float64)
                  fgmask = cv2.morphologyEx(fgmask, cv2.MORPH_CLOSE, kernel)
                  fgmask = fgmask.astype(np.float64)
                  res += (40 * fgmask + gray) * 0.01
                  res_show = res / res.max()
                  res_show = np.floor(res_show * 255)
                  res_show = res_show.astype(np.uint8)
                  res_show = cv2.applyColorMap(res_show, cv2.COLORMAP_JET)

                  image_np = cv2.resize(image_np, (480, 360))  # 사이즈 수정

                  image_np2 = copy.deepcopy(image_np) #1번화면 2번에복사

                  # 10초마다 디텍션하기
                  if (Detectflag == True or (threadtimesum) % 20 == 0):
                      Detectflag = False
                      print(threadtimesum)


                      # the array based representation of the image will be used later in order to prepare the
                      # result image with boxes and labels on it.
                      # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
                      image_np_expanded = np.expand_dims(image_np2, axis=0)
                      image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')
                      # Each box represents a part of the image where a particular object was detected.
                      boxes = detection_graph.get_tensor_by_name('detection_boxes:0')
                      # Each score represent how level of confidence for each of the objects.
                      # Score is shown on the result image, together with the class label.
                      scores = detection_graph.get_tensor_by_name('detection_scores:0')
                      classes = detection_graph.get_tensor_by_name('detection_classes:0')
                      num_detections = detection_graph.get_tensor_by_name('num_detections:0')
                      # Actual detection.
                      (boxes, scores, classes, num_detections) = sess.run(
                          [boxes, scores, classes, num_detections],
                          feed_dict={image_tensor: image_np_expanded})
                      # Visualization of the results of a detection.
                      vis_util.visualize_boxes_and_labels_on_image_array(
                          image_np2,
                          np.squeeze(boxes),
                          np.squeeze(classes).astype(np.int32),
                          np.squeeze(scores),
                          category_index,
                          use_normalized_coordinates=True,
                          line_thickness=8)

                      threshold = 0.5  # 0.5이상만 보여주기

                      ## matplot에 좌표찍고 타원그리기
                      # The following code replaces the 'print ([category_index...' statement
                      height = 360
                      width = 480

                      for index, value in enumerate(classes[0]):
                          ymin = boxes[0][index][0] * height
                          xmin = boxes[0][index][1] * width
                          ymax = boxes[0][index][2] * height
                          xmax = boxes[0][index][3] * width
                          personclassname = (category_index.get(value)).get('name')
                          widthvalue = int((xmax - xmin) / 2)  # width 길이
                          heightvalue = int((ymax - ymin) / 2)  # height 길이
                          if scores[0, index] > threshold and personclassname == 'person':
                              framesize = framesize + 1  # 프레임 크기
                              framesize2 = framesize2 + 1  # 프레임 크기
                              frame2.append(framesize2)

                              others.append(0)
                              ws['E'+str(framesize2)] = '0'  # 사람을 탐지했을 때 others에 0값 넣기.
                              sf4.scatter(frame2, others)

                              frame.append(framesize)  # 프레임 값 배열에넣기
                              x.append(xmin)  # x값 배열에 값추가
                              y.append(ymin)  # y값 배열에 추가

                              # 타원그리기
                              if (width > height):
                                  # 가로 타원형 각도
                                  degree = math.atan((heightvalue / 2) / (width / 2)) * 100
                                  degreelist.append(degree)
                                  line4.set_xdata(frame)
                                  line4.set_ydata(degreelist)
                                  ws['D' + str(framesize)] = str(degree)  # 사람을 탐지했을 때 others에 0값 넣기.
                                  image_np2 = cv2.ellipse(image_np2, (int((xmax + xmin) / 2), int((ymax + ymin) / 2)),
                                                         (heightvalue, widthvalue), 90, 0, 360, (255, 255, 0), 3)
                                  bg = cv2.circle(bg, (int((xmax + xmin) / 2), int((ymax + ymin) / 2)), 25, (255, 0, 0), -1)#사람 좌표 표시


                              else:
                                  # 세로 타원형 각도
                                  degree = math.atan((width / 2) / (heightvalue / 2))
                                  degreelist.append(degree)
                                  line4.set_xdata(frame)
                                  line4.set_ydata(degreelist)
                                  ws['D' + str(framesize)] = str(degree)  # 사람을 탐지했을 때 others에 0값 넣기.
                                  image_np2 = cv2.ellipse(image_np2, (int((xmax + xmin) / 2), int((ymax + ymin) / 2)),
                                                         (widthvalue, heightvalue), 90, 0, 360, (255, 255, 0), 3)
                                  bg = cv2.circle(bg, (int((xmax + xmin) / 2), int((ymax + ymin) / 2)), 25, (255, 0, 0), -1)#사람 좌표 표시


                              # 갱신된 값 데이터 플로팅 객체에 등록 x+y의 합산값
                              avg = xmin + ymin
                              sumlist.append(avg)
                              line1.set_xdata(frame)
                              ws['A' + str(framesize)] = str(avg)  # 사람을 탐지했을 때 others에 0값 넣기.
                              line1.set_ydata(sumlist)

                              #curs.execute(sql, (avg)) #sql쿼리

                              # 정확도 값
                              scorelist.append(scores[0, index] * 100)
                              line2.set_xdata(frame)
                              ws['B' + str(framesize)] = str(scores[0, index] * 100)  # 사람을 탐지했을 때 others에 0값 넣기.
                              line2.set_ydata(scorelist)
                              # 비율
                              ratio = (xmax - xmin) / (ymax - ymin)
                              ratiolist.append(ratio)
                              line3.set_xdata(frame)
                              ws['C' + str(framesize)] = str(ratio)  # 사람을 탐지했을 때 others에 0값 넣기.
                              line3.set_ydata(ratiolist)
                              plt.draw(), plt.pause(1)  # 그래프 그리기
                          #다른것 찍기
                          if scores[0, index] > threshold and personclassname != 'person':
                              framesize2 = framesize2 + 1  # 프레임 크기
                              frame2.append(framesize2)  # 프레임 값 배열에넣기
                              # 다른것들
                              sumother = xmin + ymin
                              ws['E' + str(framesize2)] = str(sumother)  # 사람을 탐지했을 때 others에 0값 넣기.
                              others.append(sumother)
                              bg = cv2.circle(bg, (int((xmax + xmin) / 2), int((ymax + ymin) / 2)), 25, (0, 0, 255),-1)  # 사물 좌표 표시

                              sf4.scatter(frame2, others)
                              plt.draw(), plt.pause(1)  # 그래프 그리기


                      alphabg = copy.deepcopy(bg)
                      cv2.addWeighted(backbg, 0.6, alphabg, 1 - 0.6, 0, alphabg)#그린 것 반투명처리
                      #numpy_horizontal2 = np.hstack((alphabg, res_show))
                      #cv2.imshow('pattern', numpy_horizontal2)

                      convertimg = copy.deepcopy(image_np2)  #objectdetection 적용된이미지
                      #cv2.imwrite('test_images/img'+ '.jpg', convertimg)

                  resultimg1 = hconcat_resize_min([image_np, convertimg])
                  resultimg2 = hconcat_resize_min([alphabg, res_show])
                  resultimg1 = cv2.resize(resultimg1, (960, 360))
                  resultimg2 = cv2.resize(resultimg2, (960, 360))
                  numpy_horizontal = np.vstack((resultimg1, resultimg2))

                  fig.savefig(filename+"_out_rcnninception.png") #matplot 이미지 저장
                  wb.save(filename+'_out_rcnninception.xlsx')  # 엑셀에 저장
                  out.write(numpy_horizontal)
                  cv2.imshow('object_detection', numpy_horizontal)

                  if cv2.waitKey(25) & 0xFF == ord('q'):
                      cv2.destroyAllWindows()
                      # conn.commit()
                      # conn.close()
                      break

t = Thread()
t.start()